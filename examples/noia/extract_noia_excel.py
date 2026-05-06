"""
Extrae los datos de Noia de toda la colección de .DAT del Ministerio que el
usuario tiene en `D:\\PSdeG-PSOE de Noia\\Datos Elecciones\\` y los vuelca en
un único Excel con varias pestañas.

Pestañas:
    1. Cobertura           — qué hay y qué falta en cada año
    2. Resumen             — un row por elección con totales globales de Noia
    3. Partidos por elec.  — cada candidatura con votos, escaños, % válidos
    4. Candidatos          — listas completas. 1979-1999 transcritas a mano del
                              BOP A Coruña (data/candidaturas_pre2003_noia.json);
                              2003-2023 extraídas del .DAT 04 del Ministerio.
    5. Mesa a mesa         — censo/votantes/blancos/nulos por mesa (1987+)
    6. Mesas x partido     — votos por mesa por candidatura (1987+)

Solo procesa elecciones MUNICIPALES (proceso 04).
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from infoelectoral.parser import parse_file  # noqa: E402

USER_BASE = Path(r"D:\PSdeG-PSOE de Noia\Datos Elecciones")
OUTPUT = Path(r"D:\PSdeG-PSOE de Noia\Noia_Elecciones_completas.xlsx")
PRE2003_JSON = ROOT / "data" / "candidaturas_pre2003_noia.json"
LOCALES_JSON = ROOT / "data" / "locales_electorales_noia.json"

NOIA_NAME = "Noia"
NOIA_INE = "15057"


def _load_locales() -> dict[str, dict[str, dict[str, str]]]:
    if not LOCALES_JSON.exists():
        return {}
    raw = json.loads(LOCALES_JSON.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def _mesa_key(distrito, seccion, mesa) -> str:
    """Construye la clave canónica de una mesa: 'DD-SSS-M'.

    `distrito` viene del .DAT como '01'/'02'/'03' o None (que = '01').
    `seccion` viene como '0001'/'0002'/.../'0006' (4 chars con padding).
    `mesa` es 'A'/'B'/'C'/'U' o None (que = 'U').
    """
    d = (distrito or "01").strip().zfill(2)
    s = (seccion or "").strip()
    # quitar el primer cero si la sección viene a 4 chars
    if len(s) == 4 and s.startswith("0"):
        s = s[1:]
    s = s.zfill(3)
    m = (mesa or "U").strip() or "U"
    return f"{d}-{s}-{m}"


def _lookup_local(locales: dict, year: int, distrito, seccion, mesa) -> tuple[str, str]:
    year_dict = locales.get(str(year), {})
    if not year_dict:
        return "", ""
    key = _mesa_key(distrito, seccion, mesa)
    info = year_dict.get(key)
    if info:
        return info.get("local", ""), info.get("direccion", "")

    # Fallback A↔U: en algunos años el .DAT codifica las mesas únicas como 'A'
    # mientras el BOP las publica como 'U' (o viceversa). Probamos el flip.
    if key.endswith("-A"):
        alt = key[:-2] + "-U"
    elif key.endswith("-U"):
        alt = key[:-2] + "-A"
    else:
        return "", ""
    info = year_dict.get(alt)
    if info:
        return info.get("local", ""), info.get("direccion", "")
    return "", ""

ELECTIONS = [
    # (sufijo del fichero AAMM, fecha legible)
    ("7904", "1979-04-03"),
    ("8305", "1983-05-08"),
    ("8706", "1987-06-10"),
    ("9105", "1991-05-26"),
    ("9505", "1995-05-28"),
    ("9906", "1999-06-13"),
    ("0305", "2003-05-25"),
    ("0705", "2007-05-27"),
    ("1105", "2011-05-22"),
    ("1505", "2015-05-24"),
    ("1905", "2019-05-26"),
    ("2305", "2023-05-28"),
]


def _find_dat(year_dir: Path, file_code: str, suffix: str, prefer_subdir: str | None = None) -> Path | None:
    """Busca un .DAT por código y sufijo dentro del directorio del año.
    Si prefer_subdir está dado y existe, busca primero ahí."""
    name = f"{file_code}04{suffix}.DAT"
    if prefer_subdir:
        cand = year_dir / prefer_subdir / name
        if cand.exists() and cand.stat().st_size > 0:
            return cand
    # Fallback: cualquier sitio dentro del año
    for cand in year_dir.rglob(name):
        if cand.stat().st_size > 0:
            return cand
    return None


def _build_party_catalog(file_03: Path | None) -> dict[str, dict]:
    """codigo_candidatura -> {siglas, candidatura, prov_acum, auto_acum, nac_acum}"""
    if file_03 is None:
        return {}
    res = parse_file(file_03)
    out: dict[str, dict] = {}
    for r in res.rows:
        codigo = r.get("Código")
        if codigo:
            out[codigo] = {
                "Siglas": r.get("Siglas", ""),
                "Candidatura": r.get("Candidatura", ""),
                "Cab. provincial": r.get("Candidatura provincial", ""),
                "Cab. autonómica": r.get("Candidatura autonómica", ""),
                "Cab. nacional": r.get("Candidatura nacional", ""),
            }
    return out


def _filter_noia(rows: list[dict]) -> list[dict]:
    return [r for r in rows if r.get("Municipio") == NOIA_NAME]


def _norm(s: str) -> str:
    """Normaliza un texto para comparación tolerante (mayúsculas, sin tildes, sin puntuación)."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip().upper()


# Tabla de aliases por siglas del JSON pre-2003 → patrones que deben aparecer
# en (siglas o nombre) del .DAT (ya normalizados con _norm). Construida tras
# inspeccionar las candidaturas reales que el Ministerio publica para Noia
# en los .DAT del proceso 04 entre 1979 y 1999.
PARTIDO_ALIASES: dict[str, list[str]] = {
    "PSOE":         ["PSOE", "SOCIALISTA OBRERO"],
    "PSdeG-PSOE":   ["PSDEGPSOE", "PSDG PSOE", "PSG PSOE", "PS G PSOE",
                     "PSDEG PSOE", "SOCIALISTAS DE GALICIA"],
    "BNG":          ["BLOQUE NACIONALISTA"],
    "BNPG":         ["BNPG", "BLOQUE NACIONAL POPULAR"],
    "PP":           ["PARTIDO POPULAR"],
    "AP":           ["FEDERACION DE PARTIDOS DE ALIANZA POPULAR", "ALIANZA POPULAR"],
    "AP-PDP-UL":    ["ALIANZA POPULAR PARTI", "AP PDP UL"],
    "UCD":          ["UNION DE CENTRO DEMOCRATICO"],
    "CD":           ["COALICION DEMOCRATICA"],
    "PCG":          ["PARTIDO COMUNISTA DE GALICIA"],
    "EU-EG":        ["ESQUERDA UNIDA"],
    "AG":           ["ALTERNATIVA GALEGA"],
    "CIN":          ["CANDIDATURA INDEPENDENTE", "CANDIDATURA INDEPENDIENTE"],
    "GI":           ["INDEP", "GRUPO INDEPENDIENTE"],
    "AMDG":         ["AMDG"],  # quizás no figura en el .DAT como tal
    "PTG":          ["PARTIDO DO TRABALLO"],
    "CPG":          ["COALICION PROGRESISTA"],
    "PTG-UC-PTE":   ["PART DOS TRABALLAD", "PTG UC", "TRABALLAD DE GALICIA UNIDAD COMUNISTA"],
    "CDS":          ["CENTRO DEMOCRATICO Y SOCIAL"],
    "PGI":          ["GALLEGO INDEPENDIENTE"],
    "PSG-EG":       ["PARTIDO SOCIALISTA GALEGO ESQUERDA"],
}


def _match_partido_pre2003(siglas_pre: str, nombre_pre: str, partidos_dat: dict[str, dict]) -> str | None:
    """Empareja una candidatura del JSON pre-2003 con el código del .DAT.

    Estrategia (en orden):
      1. Tabla de aliases por siglas del JSON: cada alias es un substring
         que debe aparecer en (siglas + nombre) del .DAT, normalizado.
      2. Igualdad exacta de siglas tras normalizar.
      3. Nombre del JSON contenido en el nombre del .DAT (o viceversa).
    Devuelve el código del .DAT o None.
    """
    sig_n = _norm(siglas_pre)
    nom_n = _norm(nombre_pre)

    # 1) Aliases
    aliases = PARTIDO_ALIASES.get(siglas_pre, [])
    if aliases:
        for cod, meta in partidos_dat.items():
            haystack = _norm(f"{meta.get('Siglas') or ''} {meta.get('Candidatura') or ''}")
            haystack_nospace = haystack.replace(" ", "")
            for alias in aliases:
                a_n = _norm(alias)
                a_nospace = a_n.replace(" ", "")
                if a_n in haystack or a_nospace in haystack_nospace:
                    return cod

    # 2) Igualdad exacta de siglas tras quitar espacios y puntuación
    sig_clean = sig_n.replace(" ", "")
    for cod, meta in partidos_dat.items():
        dat_sig_clean = _norm(meta.get("Siglas") or "").replace(" ", "")
        if sig_clean and sig_clean == dat_sig_clean:
            return cod

    # 3) Nombre contenido en nombre
    for cod, meta in partidos_dat.items():
        dat_nom = _norm(meta.get("Candidatura") or "")
        if nom_n and dat_nom and (nom_n in dat_nom or dat_nom in nom_n):
            return cod

    return None


def _load_pre2003() -> dict[str, list[dict]]:
    if not PRE2003_JSON.exists():
        return {}
    raw = json.loads(PRE2003_JSON.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def main() -> int:
    if not USER_BASE.exists():
        raise SystemExit(f"No existe la carpeta de datos: {USER_BASE}")

    locales = _load_locales()

    coverage_rows: list[dict] = []
    resumen_rows: list[dict] = []
    partidos_rows: list[dict] = []
    candidatos_rows: list[dict] = []
    mesas_rows: list[dict] = []
    mesas_partido_rows: list[dict] = []

    for suffix, fecha in ELECTIONS:
        year = int("19" + suffix[:2]) if int(suffix[:2]) > 70 else int("20" + suffix[:2])
        year_dir = USER_BASE / str(year)
        if not year_dir.exists():
            print(f"[{year}] No existe {year_dir}, salto.")
            continue

        # Localizamos los .DAT que necesitamos. Para 1987+ preferimos la subcarpeta
        # "Mesas" porque trae todos los ficheros incluidos los 09 y 10.
        prefer = "Mesas" if (year_dir / "Mesas").exists() else None

        f03 = _find_dat(year_dir, "03", suffix, prefer)
        f04 = _find_dat(year_dir, "04", suffix, prefer)
        f05 = _find_dat(year_dir, "05", suffix, prefer)
        f06 = _find_dat(year_dir, "06", suffix, prefer)
        f09 = _find_dat(year_dir, "09", suffix, prefer)
        f10 = _find_dat(year_dir, "10", suffix, prefer)

        partidos = _build_party_catalog(f03)

        # ---- Cobertura (la rellenaremos al final con cuántas filas Noia salieron) -

        # ---- Resumen Noia (fichero 05) -----------------------------------
        if f05:
            res05 = parse_file(f05)
            noia_05 = _filter_noia(res05.rows)
            # Puede haber varios distritos municipales — sumamos / consolidamos.
            if noia_05:
                # Tomamos el row sin distrito (Número de distrito = None) que es el agregado
                agg = next((r for r in noia_05 if r.get("Número de distrito") is None), noia_05[0])
                votantes = (agg.get("Votantes en blanco", 0) or 0) + (agg.get("Votantes nulos", 0) or 0) + (agg.get("Votantes a candidaturas", 0) or 0)
                censo_inebox = agg.get("Censo del INE") or agg.get("Censo de escrutinio")
                participacion = round(100.0 * votantes / censo_inebox, 2) if censo_inebox else None
                resumen_rows.append({
                    "Año": year,
                    "Fecha": fecha,
                    "Comunidad": agg.get("Comunidad autónoma", ""),
                    "Provincia": agg.get("Provincia", ""),
                    "Municipio": agg.get("Municipio", ""),
                    "Población de derecho": agg.get("Población de derecho"),
                    "Censo INE": agg.get("Censo del INE"),
                    "Censo escrutinio": agg.get("Censo de escrutinio"),
                    "Nº de mesas": agg.get("Número de mesas"),
                    "Votantes blancos": agg.get("Votantes en blanco"),
                    "Votantes nulos": agg.get("Votantes nulos"),
                    "Votantes a candidaturas": agg.get("Votantes a candidaturas"),
                    "Total votantes": votantes,
                    "Participación %": participacion,
                    "Nº escaños": agg.get("Número de escaños"),
                    "Datos oficiales": agg.get("Datos oficiales"),
                })

        # ---- Partidos por elección (fichero 06) --------------------------
        if f06:
            res06 = parse_file(f06)
            noia_06 = _filter_noia(res06.rows)
            # Total votos válidos para %
            total_validos = sum(r.get("Votos obtenidos", 0) or 0 for r in noia_06)
            for r in noia_06:
                codigo = r.get("Código de candidatura")
                meta = partidos.get(codigo, {})
                votos = r.get("Votos obtenidos") or 0
                partidos_rows.append({
                    "Año": year,
                    "Fecha": fecha,
                    "Provincia": r.get("Provincia"),
                    "Municipio": r.get("Municipio"),
                    "Distrito": r.get("Número de distrito"),
                    "Código": codigo,
                    "Siglas": meta.get("Siglas", ""),
                    "Candidatura": meta.get("Candidatura", ""),
                    "Votos obtenidos": votos,
                    "Concejales obtenidos": r.get("Número de candidatos obtenidos"),
                    "% sobre válidos": round(100.0 * votos / total_validos, 2) if total_validos else None,
                })

        # ---- Candidatos en Noia (fichero 04) -----------------------------
        if f04:
            res04 = parse_file(f04)
            noia_04 = _filter_noia(res04.rows)
            for r in noia_04:
                codigo = r.get("Candidatura")
                meta = partidos.get(codigo, {})
                nombre = r.get("Nombre", "") or ""
                a1 = r.get("Primer apellido", "") or ""
                a2 = r.get("Segundo apellido", "") or ""
                full = " ".join(part for part in [a1, a2] if part)
                full = f"{full}, {nombre}".strip(", ").strip()
                fecha_nac = None
                d, m, a = r.get("Día de nacimiento"), r.get("Mes de nacimiento"), r.get("Año de nacimiento")
                if d and m and a:
                    fecha_nac = f"{a:04d}-{m:02d}-{d:02d}"
                candidatos_rows.append({
                    "Año": year,
                    "Fecha": fecha,
                    "Municipio": r.get("Municipio"),
                    "Código candidatura": codigo,
                    "Siglas": meta.get("Siglas", ""),
                    "Candidatura": meta.get("Candidatura", ""),
                    "Orden": r.get("Número de orden del candidato"),
                    "Tipo": r.get("Tipo de candidato"),
                    "Apellidos, Nombre": full,
                    "Nombre": nombre,
                    "Primer apellido": a1,
                    "Segundo apellido": a2,
                    "Sexo": r.get("Sexo"),
                    "Fecha nacimiento": fecha_nac,
                    "DNI": r.get("DNI"),
                    "Elegido": r.get("Elegido"),
                })

        # ---- Mesa a mesa Noia (fichero 09) -------------------------------
        if f09:
            res09 = parse_file(f09)
            noia_09 = _filter_noia(res09.rows)
            for r in noia_09:
                votantes = (r.get("Votantes en blanco", 0) or 0) + (r.get("Votantes nulos", 0) or 0) + (r.get("Votantes a candidaturas", 0) or 0)
                censo = r.get("Censo del INE") or r.get("Censo de escrutinio")
                local, direccion = _lookup_local(
                    locales, year,
                    r.get("Número de distrito"),
                    r.get("Código de sección"),
                    r.get("Código de mesa"),
                )
                mesas_rows.append({
                    "Año": year,
                    "Fecha": fecha,
                    "Distrito": r.get("Número de distrito"),
                    "Sección": (r.get("Código de sección") or "").strip(),
                    "Mesa": r.get("Código de mesa") or "U",
                    "Local electoral": local,
                    "Dirección del local": direccion,
                    "Censo INE": r.get("Censo del INE"),
                    "Censo escrutinio": r.get("Censo de escrutinio"),
                    "Censo CERE": r.get("Censo CERE en escrutinio"),
                    "Total votantes CERE": r.get("Total votantes CERE"),
                    "1er avance participación": r.get("Votantes del primer avance"),
                    "2º avance participación": r.get("Votantes del segundo avance"),
                    "Votantes blancos": r.get("Votantes en blanco"),
                    "Votantes nulos": r.get("Votantes nulos"),
                    "Votantes a candidaturas": r.get("Votantes a candidaturas"),
                    "Total votantes": votantes,
                    "Participación %": round(100.0 * votantes / censo, 2) if censo else None,
                    "Datos oficiales": r.get("Datos oficiales"),
                })

        # ---- Mesa por partido Noia (fichero 10) --------------------------
        if f10:
            res10 = parse_file(f10)
            noia_10 = _filter_noia(res10.rows)
            for r in noia_10:
                codigo = r.get("Código de candidatura")
                meta = partidos.get(codigo, {})
                local, _direccion = _lookup_local(
                    locales, year,
                    r.get("Número de distrito"),
                    r.get("Código de sección"),
                    r.get("Código de mesa"),
                )
                mesas_partido_rows.append({
                    "Año": year,
                    "Fecha": fecha,
                    "Distrito": r.get("Número de distrito"),
                    "Sección": (r.get("Código de sección") or "").strip(),
                    "Mesa": r.get("Código de mesa") or "U",
                    "Local electoral": local,
                    "Código candidatura": codigo,
                    "Siglas": meta.get("Siglas", ""),
                    "Candidatura": meta.get("Candidatura", ""),
                    "Votos": r.get("Votos obtenidos"),
                })

        n_part = sum(1 for r in partidos_rows if r['Año']==year)
        n_cand = sum(1 for r in candidatos_rows if r['Año']==year)
        n_mesas = sum(1 for r in mesas_rows if r['Año']==year)
        n_mxp = sum(1 for r in mesas_partido_rows if r['Año']==year)

        coverage_rows.append({
            "Año": year,
            "Fecha": fecha,
            "Globales Noia (05)": "Sí" if f05 else "AUSENTE",
            "Partidos en Noia (06)": f"{n_part} candidaturas" if f06 else "AUSENTE",
            "Candidatos de Noia (04)": (
                f"{n_cand} candidatos" if n_cand
                else ("AUSENTE EN .DAT (Ministerio no publicó listas)" if not f04
                      else "VACÍO PARA NOIA (sólo capitales en este año)")
            ),
            "Mesa a mesa Noia (09)": f"{n_mesas} mesas" if n_mesas else "—",
            "Mesas × partido (10)": f"{n_mxp} filas" if n_mxp else "—",
        })

        print(
            f"[{year}] resumen={len(noia_05) if f05 else 0}  partidos={n_part}  "
            f"candidatos={n_cand}  mesas={n_mesas}  mesas_x_partido={n_mxp}"
        )

    # ---- Integración candidaturas pre-2003 (BOP A Coruña) ------------------
    pre2003 = _load_pre2003()
    if pre2003:
        # Para cruzar elegido/no, agrupamos escaños por código de candidatura del .DAT
        escanos_por_anio_y_codigo: dict[int, dict[str, int]] = defaultdict(dict)
        partidos_por_anio: dict[int, dict[str, dict]] = defaultdict(dict)
        for r in partidos_rows:
            escanos_por_anio_y_codigo[r["Año"]][r["Código"]] = r["Concejales obtenidos"] or 0
        # Reconstruimos catálogo de partidos por año a partir de partidos_rows
        for r in partidos_rows:
            partidos_por_anio[r["Año"]][r["Código"]] = {
                "Siglas": r["Siglas"], "Candidatura": r["Candidatura"],
            }

        candidatos_pre_count = 0
        for year_str, listas in pre2003.items():
            year = int(year_str)
            # Si en candidatos_rows ya hay filas para este año (no debería para 1979-1999),
            # las dejamos y no metemos las del JSON.
            if any(r["Año"] == year for r in candidatos_rows):
                continue

            partidos_dat = partidos_por_anio.get(year, {})
            escanos_dat = escanos_por_anio_y_codigo.get(year, {})

            fecha_year = next((f for s, f in ELECTIONS if int(("19" if int(s[:2]) > 70 else "20") + s[:2]) == year), "")

            for cand in listas:
                cod_dat = _match_partido_pre2003(cand["siglas"], cand["nombre"], partidos_dat)
                escanos = escanos_dat.get(cod_dat, 0) if cod_dat else 0

                # Titulares
                for orden, nombre in enumerate(cand["titulares"], start=1):
                    candidatos_rows.append({
                        "Año": year,
                        "Fecha": fecha_year,
                        "Municipio": NOIA_NAME,
                        "Código candidatura": cod_dat or "",
                        "Siglas": cand["siglas"],
                        "Candidatura": cand["nombre"],
                        "Orden": orden,
                        "Tipo": "Titular",
                        "Apellidos, Nombre": nombre,
                        "Nombre": "",
                        "Primer apellido": "",
                        "Segundo apellido": "",
                        "Sexo": "",
                        "Fecha nacimiento": "",
                        "DNI": "",
                        "Elegido": "Sí" if orden <= escanos else "No",
                    })
                    candidatos_pre_count += 1

                # Suplentes
                for orden, nombre in enumerate(cand.get("suplentes", []), start=1):
                    candidatos_rows.append({
                        "Año": year,
                        "Fecha": fecha_year,
                        "Municipio": NOIA_NAME,
                        "Código candidatura": cod_dat or "",
                        "Siglas": cand["siglas"],
                        "Candidatura": cand["nombre"],
                        "Orden": orden,
                        "Tipo": "Suplente",
                        "Apellidos, Nombre": nombre,
                        "Nombre": "",
                        "Primer apellido": "",
                        "Segundo apellido": "",
                        "Sexo": "",
                        "Fecha nacimiento": "",
                        "DNI": "",
                        "Elegido": "—",
                    })
                    candidatos_pre_count += 1

            # Actualizar la fila de cobertura del año con el conteo
            for cov in coverage_rows:
                if cov["Año"] == year:
                    cov["Candidatos de Noia (04)"] = (
                        f"{candidatos_pre_count} candidatos (BOP A Coruña, transcripción manual)"
                        if any(r["Año"] == year and r.get("Código candidatura") != "AUSENTE" for r in candidatos_rows)
                        else cov["Candidatos de Noia (04)"]
                    )
                    break

        # Recalculamos la cobertura sustituyendo el conteo por año
        for cov in coverage_rows:
            year = cov["Año"]
            n = sum(1 for r in candidatos_rows if r["Año"] == year)
            if n and str(year) in pre2003:
                cov["Candidatos de Noia (04)"] = f"{n} candidatos (BOP A Coruña, transcripción manual)"
            elif n:
                cov["Candidatos de Noia (04)"] = f"{n} candidatos"

        print(f"\nCandidaturas pre-2003 integradas: {candidatos_pre_count} filas añadidas a 'Candidatos'.")

    # Normalizamos el campo "Tipo" basándonos en el orden y el número de escaños
    # del municipio (Noia tiene 17 escaños desde 1979). El campo del .DAT a veces
    # viene mal codificado (p. ej. en 2007 y 2011 el Ministerio marcó toda la lista
    # del PSOE como "Suplente" siendo claramente titulares). Reglas:
    #   - orden ∈ [1, 17]  → Titular
    #   - orden > 17       → Suplente
    # (esto vale para municipios de 5.001 a 50.000 habitantes; Noia ha estado
    # siempre en ese tramo durante el periodo cubierto.)
    NOIA_ESCANOS = 17
    for r in candidatos_rows:
        orden = r.get("Orden") or 0
        if orden and orden <= NOIA_ESCANOS:
            r["Tipo"] = "Titular"
        elif orden:
            r["Tipo"] = "Suplente"

    # Reordenamos candidatos_rows por (Año, Siglas, Tipo, Orden)
    tipo_order = {"Titular": 0, "Suplente": 1, "": 2}
    candidatos_rows.sort(key=lambda r: (
        r["Año"], r.get("Siglas", ""),
        tipo_order.get(r.get("Tipo", ""), 9),
        r.get("Orden") or 0,
    ))

    # ---- Escritura del Excel -----------------------------------------------
    print(f"\nEscribiendo Excel en {OUTPUT}…")
    write_excel(
        OUTPUT,
        {
            "Cobertura": coverage_rows,
            "Resumen": resumen_rows,
            "Partidos por elección": partidos_rows,
            "Candidatos": candidatos_rows,
            "Mesa a mesa": mesas_rows,
            "Mesas x partido": mesas_partido_rows,
        },
    )
    print("OK")
    return 0


def write_excel(path: Path, sheets: dict[str, list[dict]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Quitamos la hoja por defecto
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            ws.append(["(sin datos)"])
            continue

        # Cabeceras
        cols = list(rows[0].keys())
        ws.append(cols)
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Datos
        for r in rows:
            ws.append([r.get(c) for c in cols])

        # Freeze header + autofilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Anchura de columnas en función del contenido (cap a 40)
        for col_idx, name in enumerate(cols, 1):
            max_len = len(str(name))
            for r in rows[:200]:  # muestreamos 200 filas para no perder tiempo
                v = r.get(name)
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 42)

    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
